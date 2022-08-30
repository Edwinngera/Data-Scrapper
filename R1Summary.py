import openpyxl as xl


# opening the source excel file
filename = "Combined R1.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]




# opening the destination excel file
filename1 = "211005 Candidates Summary.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

#candidates Name
for cell in ws1['B:B']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=1, value=cell.value)

#English Scores
for cell in ws1['F:F']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=10, value=cell.value)

#Presence
for cell in ws1['G:G']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=11, value=cell.value)

#Logical Thinking
for cell in ws1['H:H']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=12, value=cell.value)

#Overall 
for cell in ws1['I:I']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=13, value=cell.value)

#Notes
for cell in ws1['J:J']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=40, value=cell.value)

#Interviewer 1
for cell in ws1['AQ:AQ']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=9, value=cell.value)



"""Interview 2 """
#Candidates Name
for cell in ws1['Q:Q']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=14, value=cell.value)

#English Scores
for cell in ws1['R:R']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=15, value=cell.value)

#Presence
for cell in ws1['S:S']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=16, value=cell.value)

#Logical Thinking
for cell in ws1['T:T']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=17, value=cell.value)


#Overall
for cell in ws1['U:U']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=18, value=cell.value)



#Notes
for cell in ws1['V:V']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=41, value=cell.value)


#Interviewer 2 Name
for cell in ws1['BC:BC']:
    if(cell.row > 3):
      ws2.cell(row=cell.row, column=14, value=cell.value)





wb2.save(str(filename1))
