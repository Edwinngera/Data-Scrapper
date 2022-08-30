import openpyxl as xl
import pylightxl as xz


# opening the source excel file
filename = "Combined R2.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

db = xz.readxl(filename)


# opening the destination excel file
filename1 = "211005 Candidates Summary.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

db1=xz.readxl(filename1)

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

#R2 marks Interviewer one
for cell in ws1['B:B']:
    if(cell.row > 5): 
        for cell2 in ws2['A:A']:
              if(cell2.row > 3): 
                if(cell2.value!=None and cell.value!=None):
                  if(cell2.value.strip().lower()==cell.value.strip().lower()):
                      print(cell.value.strip())
                      array_values=db.ws(ws='Sheet1').row(row=cell.row)
                      column=21
                      marks=array_values[8:17]
                      for i in range(0,len(marks)):
                          #db1.ws(ws='Sheet1').update_index(row=cell2.row, col=column, val=marks[i])
                          ws2.cell(row=cell2.row, column=column, value=marks[i])
                          column+=1

#R2 Marks Interviewer2
for cell in ws1['B:B']:
    if(cell.row > 5): 
        for cell2 in ws2['A:A']:
              if(cell2.row > 3):
                if(cell2.value!=None and cell.value!=None):
                  if(cell2.value.strip().lower()==cell.value.strip().lower()):
                      array_values=db.ws(ws='Sheet1').row(row=cell.row)
                      column=31
                      marks=array_values[33:42]
                      for i in range(0,len(marks)):
                          #db1.ws(ws='Sheet1').update_index(row=cell2.row, col=column, val=marks[i])
                          ws2.cell(row=cell2.row, column=column, value=marks[i])
                          column+=1


#Interviewers
#Interviewer 1 Name
for cell in ws1['B:B']:
    if(cell.row > 5): 
        for cell2 in ws2['A:A']:
              if(cell2.row > 3): 
                 if(cell2.value!=None and cell.value!=None):
                  if(cell2.value.strip().lower()==cell.value.strip().lower()):
                      print(cell2.value)
                      array_values=db.ws(ws='Sheet1').row(row=cell.row)
                      column=20
                      marks=array_values[22:23]
                      for i in range(0,len(marks)):
                          #db1.ws(ws='Sheet1').update_index(row=cell2.row, col=column, val=marks[i])
                          ws2.cell(row=cell2.row, column=column, value=marks[i])
                          column+=1

#Interviwer 2 Name
for cell in ws1['B:B']:
    if(cell.row > 5): 
        for cell2 in ws2['A:A']:
              if(cell2.row > 3): 
                if(cell2.value!=None and cell.value!=None):
                  if(cell2.value.strip().lower()==cell.value.strip().lower()):
                      print(cell2.value)
                      array_values=db.ws(ws='Sheet1').row(row=cell.row)
                      column=30
                      marks=array_values[47:48]
                      for i in range(0,len(marks)):
                          #db1.ws(ws='Sheet1').update_index(row=cell2.row, col=column, val=marks[i])
                          ws2.cell(row=cell2.row, column=column, value=marks[i])
                          column+=1



#Notes
#R2 Comments Interviwer 1
for cell in ws1['B:B']:
    if(cell.row > 5): 
        for cell2 in ws2['A:A']:
              if(cell2.row > 3): 
                if(cell2.value!=""):
                   if(cell2.value!=None and cell.value!=None):
                      array_values=db.ws(ws='Sheet1').row(row=cell.row)
                      column=42
                      marks=array_values[17:18]
                      for i in range(0,len(marks)):
                          #db1.ws(ws='Sheet1').update_index(row=cell2.row, col=column, val=marks[i])
                          ws2.cell(row=cell2.row, column=column, value=marks[i])
                          column+=1

#R2 comments Interviewer 2
for cell in ws1['B:B']:
    if(cell.row > 5): 
        for cell2 in ws2['A:A']:
              if(cell2.row > 3): 
                if(cell2.value!=None and cell.value!=None):
                  if(cell2.value.strip().lower()==cell.value.strip().lower()):
                      print(cell2.value)
                      array_values=db.ws(ws='Sheet1').row(row=cell.row)
                      column=43
                      marks=array_values[42:43]
                      for i in range(0,len(marks)):
                          #db1.ws(ws='Sheet1').update_index(row=cell2.row, col=column, val=marks[i])
                          ws2.cell(row=cell2.row, column=column, value=marks[i])
                          column+=1



wb2.save(str(filename1))
                          


                
                 






        
      



